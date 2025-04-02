import face_recognition
import cv2
import numpy as np
import os
import openpyxl
from datetime import datetime

# Initialize webcam
video_capture = cv2.VideoCapture(0)
if not video_capture.isOpened():
    print("Error: Could not open webcam.")
    exit()

# Load a sample picture and learn how to recognize it.
image_path = r"C:\Users\Dell\OneDrive\Documents\ananyaa.jpg"  # Correct image path
person_name = "Ananyaa"

# Load known image
known_image = face_recognition.load_image_file(image_path)
known_encoding = face_recognition.face_encodings(known_image)

# Ensure at least one face encoding is found
if len(known_encoding) > 0:
    known_face_encoding = known_encoding[0]
else:
    print("Error: No face detected in the known image.")
    video_capture.release()
    exit()

# List of known faces
known_face_encodings = [known_face_encoding]
known_face_names = [person_name]

# Initialize attendance sheet
lecture_name = input("Please enter the lecture name: ").strip()

# Define the correct file path for the Excel file
excel_filename = r"C:\Users\Dell\OneDrive\Documents\attendence_excel.xlsx"

# Check if the file exists, else create it
if not os.path.exists(excel_filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Records"
    wb.save(excel_filename)
    print(f"New Excel file created at {excel_filename}")

# Open the Excel file
wb = openpyxl.load_workbook(excel_filename)

# Create a new sheet for the lecture if not present
if lecture_name not in wb.sheetnames:
    sheet = wb.create_sheet(title=lecture_name)
    sheet.append(["Name", "Timestamp"])
else:
    sheet = wb[lecture_name]

# Attendance tracking
already_marked = set()

# Create a named window before entering the loop
cv2.namedWindow("Video", cv2.WINDOW_NORMAL)

while True:
    ret, frame = video_capture.read()
    if not ret:
        print("Error: Could not read frame.")
        break

    # Resize for faster processing
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

    # Detect faces and get encodings
    face_locations = face_recognition.face_locations(rgb_small_frame)
    face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

    face_names = []
    for face_encoding in face_encodings:
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        name = "Unknown"

        if True in matches:
            best_match_index = np.argmin(face_recognition.face_distance(known_face_encodings, face_encoding))
            name = known_face_names[best_match_index]

        face_names.append(name)

        # Mark attendance if not already marked
        if name != "Unknown" and name not in already_marked:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            sheet.append([name, timestamp])
            wb.save(excel_filename)  # Ensure data is saved after each update
            already_marked.add(name)
            print(f"Attendance marked for {name} at {timestamp}")

    # Display video feed with face recognition
    for (top, right, bottom, left), name in zip(face_locations, face_names):
        top, right, bottom, left = top * 4, right * 4, bottom * 4, left * 4

        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
        cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 255, 0), cv2.FILLED)
        cv2.putText(frame, name, (left + 6, bottom - 6), cv2.FONT_HERSHEY_DUPLEX, 1, (255, 255, 255), 1)

    cv2.imshow('Video', frame)

    # Exit when 'q' is pressed
    if cv2.waitKey(1) & 0xFF == ord('q'):
        print("Exiting and saving attendance.")
        break

# Cleanup
video_capture.release()
cv2.destroyAllWindows()
wb.save(excel_filename)
wb.close()  # Ensure the workbook is closed properly
print(f"Attendance successfully saved in {excel_filename}")
